import openpyxl

def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row

def get_column_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column

def get_cell_data(path, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=column_number).value

def set_cell_data(path,sheet_name,row_number,column_number,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(path)

def get_cate_from_excel(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column


    for row in range(2, total_rows + 1):
        now_list = []
        for c in range(1, total_cols + 1):
            now_list.append(sheet.cell(row=row, column=c).value)
        final_list.append(now_list)
    return final_list
